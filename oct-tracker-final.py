import pandas as pd
import os
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def find_latest_excel_files(folder_path):
    """Find the most recent left and right eye Excel files."""
    left_files = glob.glob(os.path.join(folder_path, "left_results_*.xlsx"))
    right_files = glob.glob(os.path.join(folder_path, "right_results_*.xlsx"))
    
    # Sort by modification time, newest first
    left_files.sort(key=os.path.getmtime, reverse=True)
    right_files.sort(key=os.path.getmtime, reverse=True)
    
    left_file = left_files[0] if left_files else None
    right_file = right_files[0] if right_files else None
    
    return left_file, right_file

def load_excel_data(file_path):
    """Load data from an Excel file."""
    if not file_path or not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        return None
    
    try:
        df = pd.read_excel(file_path)
        print(f"Successfully loaded data from {file_path}")
        print(f"Columns found: {df.columns.tolist()}")
        print(f"Number of rows: {len(df)}")
        
        # Ensure the dataframe has the expected columns
        required_columns = ['Date', 'Number']
        if not all(col in df.columns for col in required_columns):
            print(f"Warning: Excel file {file_path} does not have the expected columns: {required_columns}")
            return None
        
        # Convert dates to datetime objects for easier comparison
        try:
            # Try to convert with multiple formats
            for format_str in ['%d-%b-%y', '%m/%d/%y', '%Y-%m-%d']:
                try:
                    df['datetime'] = pd.to_datetime(df['Date'], format=format_str, errors='coerce')
                    if not df['datetime'].isna().all():
                        print(f"Successfully parsed dates with format: {format_str}")
                        break
                except:
                    continue
            
            # If that didn't work, try a more flexible approach
            if 'datetime' not in df.columns or df['datetime'].isna().all():
                df['datetime'] = pd.to_datetime(df['Date'], errors='coerce')
                print("Used flexible date parsing")
            
            # Sort by date
            df = df.sort_values('datetime')
            return df
        except Exception as e:
            print(f"Error converting dates: {e}")
            return df
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return None

def get_number_for_date(df, target_date):
    """Find the OCT number that corresponds to a specific date."""
    if df is None or df.empty:
        print("No data available for date lookup")
        return "999"  # Return placeholder if no data
    
    # Convert target date to datetime for comparison
    try:
        target_datetime = pd.to_datetime(target_date)
        print(f"Looking for OCT number for date: {target_date} ({target_datetime})")
    except:
        print(f"Could not parse date: {target_date}")
        return "999"
    
    # Look for exact match first
    exact_matches = df[df['datetime'] == target_datetime]
    if not exact_matches.empty:
        result = str(exact_matches.iloc[0]['Number'])
        print(f"Found exact match: {result}")
        return result
    
    # If no exact match, find the closest date (preferring dates before the target)
    df['days_diff'] = (df['datetime'] - target_datetime).dt.days
    
    # First try to find the closest date on or before the target date
    before_dates = df[df['days_diff'] <= 0]
    if not before_dates.empty:
        closest_before = before_dates.loc[before_dates['days_diff'].idxmax()]
        result = str(closest_before['Number'])
        print(f"Found closest date before: {closest_before['Date']} with OCT: {result}")
        return result
    
    # If no dates before target, find the closest date after
    after_dates = df[df['days_diff'] > 0]
    if not after_dates.empty:
        closest_after = after_dates.loc[after_dates['days_diff'].idxmin()]
        result = str(closest_after['Number'])
        print(f"Found closest date after: {closest_after['Date']} with OCT: {result}")
        return result
    
    # If we get here, no match found at all
    print("No matching OCT value found")
    return "999"

def find_lowest_between_dates(df, start_date, end_date):
    """Find the lowest OCT number between two dates, and its corresponding date."""
    if df is None or df.empty:
        return "31-Dec-99", "999"
    
    try:
        start_datetime = pd.to_datetime(start_date)
        end_datetime = pd.to_datetime(end_date)
        print(f"Looking for lowest OCT between {start_date} and {end_date}")
    except:
        print(f"Could not parse dates: {start_date} or {end_date}")
        return "31-Dec-99", "999"
    
    # Find all readings strictly between the two dates
    between_dates = df[(df['datetime'] > start_datetime) & (df['datetime'] < end_datetime)]
    
    if between_dates.empty:
        print("No OCT readings found between these dates")
        return "31-Dec-99", "999"
    
    # Find the row with the minimum number
    try:
        between_dates['Number_numeric'] = pd.to_numeric(between_dates['Number'], errors='coerce')
        min_idx = between_dates['Number_numeric'].idxmin()
        min_row = between_dates.loc[min_idx]
        min_date = min_row['Date']
        min_number = str(min_row['Number'])
        print(f"Found lowest OCT: {min_number} on date: {min_date}")
        return min_date, min_number
    except Exception as e:
        print(f"Error finding minimum: {e}")
        return "31-Dec-99", "999"

def format_date(date_str):
    """Convert date to DD-MMM-YY format."""
    try:
        date_obj = pd.to_datetime(date_str)
        return date_obj.strftime("%d-%b-%y")
    except:
        return date_str

#
# ======= REPLACED create_tracking_sheet FUNCTION STARTS HERE =======
#
def create_tracking_sheet(main_folder, patient_id, eye_code, injection_dates, final_checkup_date):
    """Create a tracking sheet with proper date formatting and correct column names."""
    # Find the most recent Excel files
    left_file, right_file = find_latest_excel_files(main_folder)
    
    # Determine which eye to use
    eye = "left" if eye_code == 0 else "right"
    excel_file = left_file if eye == "left" else right_file
    
    if not excel_file:
        print(f"No {eye} eye data found.")
        return
    
    print(f"Using data from: {excel_file}")
    
    # Load the data
    df = load_excel_data(excel_file)
    if df is None:
        print(f"Could not load data from {excel_file}")
        return
    
    # Create a new Excel workbook
    output_file = os.path.join(
        main_folder,
        f"{patient_id}_{eye}_tracking_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    )
    
    # Prepare data for output - with the exact column names requested
    data = {"Patient_number": [patient_id], "Eye": [eye]}
    column_order = ["Patient_number", "Eye"]
    
    # Process injection dates
    all_valid_dates = []  # Keep track of all valid dates for best date calculation
    current_injections = len(injection_dates)  # Initialize current_injections here
    
    for i, inj_date in enumerate(injection_dates, 1):
        # Format the date properly
        formatted_date = format_date(inj_date)
        all_valid_dates.append(inj_date)
        
        # Get OCT value for this date
        oct_value = get_number_for_date(df, inj_date)
        
        # Add date and OCT
        date_col = f"@{i}_Date_Injection"
        oct_col = f"@{i}_CFT"
        
        data[date_col] = [formatted_date]
        data[oct_col] = [oct_value]
        
        column_order.extend([date_col, oct_col])
        
        # If not the last date, add lowest between data
        if i < len(injection_dates):
            next_date = injection_dates[i]
            low_date, low_oct = find_lowest_between_dates(df, inj_date, next_date)
            
            # Format the lowest date
            formatted_low_date = format_date(low_date)
            if low_date != "31-Dec-99":
                all_valid_dates.append(low_date)
            
            # Add to data with the requested column names
            best_oct_col = f"@{i}_Best_CFT"
            best_date_col = f"@{i}_Date_Best_CFT"
            
            data[best_oct_col] = [low_oct]
            data[best_date_col] = [formatted_low_date]
            
            column_order.extend([best_oct_col, best_date_col])
        
        # If this is the last injection date and we have a final checkup, add lowest between
        if i == len(injection_dates) and final_checkup_date:
            low_date, low_oct = find_lowest_between_dates(df, inj_date, final_checkup_date)
            
            # Format the lowest date
            formatted_low_date = format_date(low_date)
            if low_date != "31-Dec-99":
                all_valid_dates.append(low_date)
            
            # Add to data with the requested column names
            best_oct_col = f"@{i}_Best_CFT"
            best_date_col = f"@{i}_Date_Best_CFT"
            
            data[best_oct_col] = [low_oct]
            data[best_date_col] = [formatted_low_date]
            
            column_order.extend([best_oct_col, best_date_col])
    
    # Fill remaining slots with placeholders
    max_injections = 8  # Maximum number of injection dates
    
    for i in range(current_injections + 1, max_injections + 1):
        # Add date and OCT columns with placeholders
        date_col = f"@{i}_Date_Injection"
        oct_col = f"@{i}_CFT"
        
        data[date_col] = ["31-Dec-99"]
        data[oct_col] = ["999"]
        
        column_order.extend([date_col, oct_col])
        
        # Add best values columns with placeholders
        if i < max_injections:  # Don't add after the last injection
            best_oct_col = f"@{i}_Best_CFT"
            best_date_col = f"@{i}_Date_Best_CFT"
            
            data[best_oct_col] = ["999"]
            data[best_date_col] = ["31-Dec-99"]
            
            column_order.extend([best_oct_col, best_date_col])
    
    # Add final visit date
    if final_checkup_date:
        formatted_final_date = format_date(final_checkup_date)
        all_valid_dates.append(final_checkup_date)
        data["Final_visit_Date"] = [formatted_final_date]
    else:
        data["Final_visit_Date"] = ["31-Dec-99"]
    
    column_order.append("Final_visit_Date")
    
    # Create DataFrame with proper column order
    output_df = pd.DataFrame(data)
    output_df = output_df[column_order]
    
    # Save to Excel
    output_df.to_excel(output_file, index=False)
    
    # Apply formatting
    apply_excel_formatting_new(output_file)
    
    print(f"Created tracking sheet: {output_file}")
    return output_file
#
# ======= REPLACED create_tracking_sheet FUNCTION ENDS HERE =======
#

def apply_excel_formatting_new(output_file):
    """Apply professional formatting to the Excel file with new column structure."""
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Silver
        date_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")    # Light gray
        oct_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")     # White
        best_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")    # Medium gray
        
        header_font = Font(bold=True, size=12)
        regular_font = Font(size=11)
        
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply formatting to header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        
        # Auto-adjust column widths first
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply formatting to data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for i, cell in enumerate(row):
                cell.border = border
                cell.alignment = center_alignment
                cell.font = regular_font
                
                # Apply different formatting based on column type
                header = ws.cell(row=1, column=cell.column).value
                if header:
                    if "Date_Injection" in header or header == "Final_visit_Date":
                        cell.fill = date_fill
                    elif "CFT" in header and "Best" not in header:
                        cell.fill = oct_fill
                    elif "Best_CFT" in header or "Date_Best_CFT" in header:
                        cell.fill = best_fill
        
        # Add some freeze panes to make it more user-friendly
        ws.freeze_panes = 'C2'
        
        # Save the workbook
        wb.save(output_file)
        print(f"Applied formatting to {output_file}")
    except Exception as e:
        print(f"Error applying Excel formatting: {e}")

def main():
    # Folder path containing your Excel files
    main_folder = input("Enter folder path containing Excel files: ")
    if not os.path.exists(main_folder):
        print(f"Error: Folder path {main_folder} does not exist.")
        # Ask if user wants to create the folder
        create_folder = input(f"Would you like to create the folder {main_folder}? (y/n): ").lower() == 'y'
        if create_folder:
            try:
                os.makedirs(main_folder)
                print(f"Created folder: {main_folder}")
            except Exception as e:
                print(f"Error creating folder: {e}")
                return
        else:
            return
    
    print("\n" + "="*60)
    print("OCT DATA TRACKER - IMPROVED VERSION")
    print("="*60)
    
    # Get patient ID
    patient_id = input("Enter patient ID: ")
    
    # Get eye code (0 for left, 1 for right)
    while True:
        eye_input = input("Enter eye (0 for left, 1 for right): ")
        if eye_input in ['0', '1']:
            eye_code = int(eye_input)
            break
        print("Invalid input. Please enter 0 for left eye or 1 for right eye.")
    
    # Get number of injection dates
    while True:
        try:
            num_injections = int(input("Enter number of injection dates (1-8): "))
            if 1 <= num_injections <= 8:
                break
            print("Please enter a number between 1 and 8.")
        except ValueError:
            print("Please enter a valid number.")
    
    # Get injection dates
    injection_dates = []
    print("\nEnter injection dates in MM/DD/YY format (e.g., 4/16/24):")
    
    for i in range(num_injections):
        while True:
            date_input = input(f"Injection date {i+1}: ")
            try:
                # Validate date format - try both with and without leading zeros
                try:
                    date_obj = datetime.strptime(date_input, "%m/%d/%y")
                except ValueError:
                    # Try alternative format with leading zeros
                    date_parts = date_input.split('/')
                    if len(date_parts) == 3:
                        month = date_parts[0].zfill(2)
                        day = date_parts[1].zfill(2)
                        year = date_parts[2].zfill(2)
                        date_obj = datetime.strptime(f"{month}/{day}/{year}", "%m/%d/%y")
                
                injection_dates.append(date_input)
                break
            except ValueError:
                print("Invalid date format. Please use MM/DD/YY format (e.g., 4/16/24).")
    
    # Get final checkup date
    final_checkup_date = None
    has_final_checkup = input("\nDo you have a final checkup date? (y/n): ").lower() == 'y'
    
    if has_final_checkup:
        while True:
            date_input = input("Final checkup date (MM/DD/YY): ")
            try:
                # Validate date format with same flexibility as above
                try:
                    date_obj = datetime.strptime(date_input, "%m/%d/%y")
                except ValueError:
                    # Try alternative format with leading zeros
                    date_parts = date_input.split('/')
                    if len(date_parts) == 3:
                        month = date_parts[0].zfill(2)
                        day = date_parts[1].zfill(2)
                        year = date_parts[2].zfill(2)
                        date_obj = datetime.strptime(f"{month}/{day}/{year}", "%m/%d/%y")
                
                final_checkup_date = date_input
                break
            except ValueError:
                print("Invalid date format. Please use MM/DD/YY format (e.g., 4/16/24).")
    
    # Create tracking sheet
    output_file = create_tracking_sheet(main_folder, patient_id, eye_code, injection_dates, final_checkup_date)
    
    if output_file:
        print(f"\nTracking sheet created: {output_file}")
        print("\nComplete! The Excel file has been formatted to match the new column structure.")
    else:
        print("\nError creating tracking sheet. Please check your inputs and try again.")

if __name__ == '__main__':
    main()
